import openpyxl
import logging
import csv
import pprint
from openpyxl import load_workbook
from openpyxl import Workbook
import glob

import requests
from bs4 import BeautifulSoup



class LotteryDataAcquisition:
# """A simple example class"""         # 三重クォートによるコメント

    def __init__(self):                  # コンストラクタ
        self.name = ""

    def create_books(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'test_sheet_1'
        # シートを追加
        wb.create_sheet()
        wb.create_sheet(index=1,title="555555")
        # 保存する
        wb.save('C:\\Users\\tamur\\Desktop\\lottery_notes\\test2.xlsx')
        glob.glob("*.xlsx")

    def create_book(self):
        wb = Workbook()
        # ワークブックの読み込み\
        wb.save('C:\\Users\\tamur\\Desktop\\lottery_notes\\myworkbook.xlsx')
        # ワークシートの選択
        ws = wb['Sheet']  # ワークシートを指定
        ws = wb.active  # アクティブなワークシートを選択
        print(f'sheet name: {ws.title}')  # sheet name: Sheet
        for sheet in wb:
            print(f'sheet name: {sheet.title}')

    # ブックデータ取得メソッド
    def lottery_notes(self):
        # ブックを取得
        get_book = openpyxl.load_workbook('C:\\Users\\tamur\\Desktop\\lottery_notes\\lottery_notes.xlsx')
        # シートを取得
        get_seat = get_book["Sheet1"]
        # セルを１つずつ取得
        for cell1 in tuple(get_seat.rows):
            for cell2 in cell1:
                print(cell2.value)

    # error logging作成
    def create_log(self):
        logging.basicConfig()
        logger = logging.getLogger(__name__)
        logger.error("エラーが発生しました")

    def hello(self):
        print("hello world55555")



    def scraping(self):
        TARGET_URL = "http://www.ohtashp.com/topics/takarakuji/index_loto6.html" # スクレイピングしたいURLをここに書く
        html = requests.get(TARGET_URL)  # HTMLを取ってくる
        soup = BeautifulSoup(html.content, "html.parser")  # HTMLを解析する
        title = soup.find("title")  # データを抽出する
        print(title)




 # クラスのインスタンスを生成
rotteryData =  LotteryDataAcquisition()
rotteryData.create_log()

rotteryData.lottery_notes()

rotteryData.create_books()
rotteryData.create_book()

rotteryData.hello()

rotteryData.scraping()

rotteryData.hello()

rotteryData.hello()
rotteryData.hello()